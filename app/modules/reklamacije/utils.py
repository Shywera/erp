"""PDF i Excel export za Reklamacije — port iz Django utils.py."""
import io
from datetime import date

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Font ──────────────────────────────────────────────────────────────────────
try:
    pdfmetrics.registerFont(TTFont("F",  "C:/Windows/Fonts/arial.ttf"))
    pdfmetrics.registerFont(TTFont("FB", "C:/Windows/Fonts/arialbd.ttf"))
    pdfmetrics.registerFont(TTFont("FI", "C:/Windows/Fonts/ariali.ttf"))
    pdfmetrics.registerFontFamily("F", normal="F", bold="FB", italic="FI", boldItalic="FB")
    _FONT = "F"
    _FONTB = "FB"
except Exception:
    _FONT = "Helvetica"
    _FONTB = "Helvetica-Bold"

# ── Boje ──────────────────────────────────────────────────────────────────────
_TAMNO   = colors.HexColor("#0F172A")
_PLAVA   = colors.HexColor("#1D4ED8")
_SIVA_TL = colors.HexColor("#475569")
_SIVA_LN = colors.HexColor("#E2E8F0")
_SIVA_BG = colors.HexColor("#F8FAFC")
_BIJELA  = colors.white

_STATUS_BOJE = {
    "NOVO":      (colors.HexColor("#DBEAFE"), colors.HexColor("#1D4ED8")),
    "U_OBRADI":  (colors.HexColor("#FEF3C7"), colors.HexColor("#92400E")),
    "CEKA":      (colors.HexColor("#CFFAFE"), colors.HexColor("#155E75")),
    "RIJESENO":  (colors.HexColor("#D1FAE5"), colors.HexColor("#065F46")),
    "ZATVORENO": (colors.HexColor("#F1F5F9"), colors.HexColor("#475569")),
}
_PRIO_BOJE = {
    "NIZAK":    colors.HexColor("#10B981"),
    "SREDNJI":  colors.HexColor("#3B82F6"),
    "VISOK":    colors.HexColor("#F59E0B"),
    "KRITICAN": colors.HexColor("#EF4444"),
}


def generiraj_pdf(r) -> io.BytesIO:
    buffer = io.BytesIO()
    PAGE_W, PAGE_H = A4
    ML = MR = 1.8 * cm
    MT = 1.6 * cm
    MB = 2.0 * cm
    CW = PAGE_W - ML - MR

    def _x(v):
        if not v:
            return "—"
        return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def zaglavlje_podnozje(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_PLAVA)
        canvas.rect(ML, PAGE_H - MT + 2 * mm, CW, 3 * mm, fill=1, stroke=0)
        canvas.setFont(_FONTB, 9)
        canvas.setFillColor(_TAMNO)
        canvas.drawString(ML, PAGE_H - MT - 4 * mm, "DEMO TISAK d.o.o.")
        canvas.setFont(_FONT, 7.5)
        canvas.setFillColor(_SIVA_TL)
        canvas.drawString(ML, PAGE_H - MT - 8.5 * mm, "Sustav upravljanja kvalitetom")
        canvas.setFont(_FONTB, 11)
        canvas.setFillColor(_PLAVA)
        canvas.drawRightString(ML + CW, PAGE_H - MT - 4 * mm, r.broj_predmeta)
        canvas.setFont(_FONT, 7.5)
        canvas.setFillColor(_SIVA_TL)
        canvas.drawRightString(ML + CW, PAGE_H - MT - 8.5 * mm,
                               f"Datum: {r.datum_prijave.strftime('%d.%m.%Y')}")
        canvas.setStrokeColor(_SIVA_LN)
        canvas.setLineWidth(0.5)
        canvas.line(ML, PAGE_H - MT - 11 * mm, ML + CW, PAGE_H - MT - 11 * mm)
        canvas.line(ML, MB - 3 * mm, ML + CW, MB - 3 * mm)
        canvas.setFont(_FONT, 7)
        canvas.setFillColor(_SIVA_TL)
        canvas.drawString(ML, MB - 7 * mm,
                          f"Generirano: {date.today().strftime('%d.%m.%Y')}  |  ERP Reklamacije")
        canvas.drawRightString(ML + CW, MB - 7 * mm, f"Stranica {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=MT + 1.5 * cm, bottomMargin=MB + 0.6 * cm,
        leftMargin=ML, rightMargin=MR,
    )

    _sC = ParagraphStyle("sC", fontName=_FONT, leading=14)

    def cel(label, value, bold=False):
        fn = _FONTB if bold else _FONT
        v = _x(value).replace("\n", "<br/>")
        return Paragraph(
            f'<font name="{_FONTB}" size="6.5" color="#64748B">{_x(label)}</font>'
            f'<br/><font name="{fn}" size="9" color="#0F172A">{v}</font>',
            _sC,
        )

    def sek(tekst):
        t = Table([[Paragraph(tekst, ParagraphStyle("SH", fontName=_FONTB,
                                                     fontSize=8, textColor=_BIJELA, leading=11))]],
                  colWidths=[CW])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _TAMNO),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ]))
        return t

    CP = [
        ("BOX", (0, 0), (-1, -1), 0.5, _SIVA_LN),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, _SIVA_LN),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]

    story = []

    # Naslov
    vrsta_bg, vrsta_fg = _STATUS_BOJE.get(r.status, (_SIVA_BG, _SIVA_TL))
    prio_color = _PRIO_BOJE.get(r.prioritet, _SIVA_TL)

    story.append(Paragraph("ZAPISNIK O NESUKLADNOSTI / REKLAMACIJI",
                            ParagraphStyle("TIT", fontName=_FONTB, fontSize=13,
                                           textColor=_TAMNO, leading=16, spaceAfter=3)))
    story.append(Paragraph(_x(r.naslov),
                            ParagraphStyle("SUB", fontName=_FONT, fontSize=10.5,
                                           textColor=_SIVA_TL, leading=14, spaceAfter=5)))

    bs = Paragraph(f" {_x(r.status_display)} ",
                   ParagraphStyle("BS", fontName=_FONTB, fontSize=8,
                                  textColor=vrsta_fg, backColor=vrsta_bg, borderPadding=(3, 6, 3, 6)))
    bp = Paragraph(f" {_x(r.prioritet_display)} ",
                   ParagraphStyle("BP", fontName=_FONTB, fontSize=8,
                                  textColor=_BIJELA, backColor=prio_color, borderPadding=(3, 6, 3, 6)))
    bv = Paragraph(f" {_x(r.vrsta_display)} ",
                   ParagraphStyle("BV", fontName=_FONT, fontSize=8,
                                  textColor=_SIVA_TL, backColor=_SIVA_BG, borderPadding=(3, 6, 3, 6)))
    badge_t = Table([[bs, bp, bv, ""]], colWidths=[3.0*cm, 2.5*cm, 4.5*cm, CW - 10.0*cm])
    badge_t.setStyle(TableStyle([("TOPPADDING", (0,0),(-1,-1),0),
                                  ("BOTTOMPADDING",(0,0),(-1,-1),8),
                                  ("LEFTPADDING",(0,0),(-1,-1),0)]))
    story.append(badge_t)

    # Identifikacija
    story.append(sek("IDENTIFIKACIJA PREDMETA"))
    col6 = CW / 6
    rok = r.rok_rjesavanja.strftime("%d.%m.%Y") if r.rok_rjesavanja else None
    kat = r.kategorija_display if r.kategorija else None

    id_rows = [
        [cel("BROJ PREDMETA", r.broj_predmeta, bold=True), "",
         cel("DATUM PRIJAVE", r.datum_prijave.strftime("%d.%m.%Y  %H:%M")), "",
         cel("VRSTA NESUKLADNOSTI", r.vrsta_display), ""],
        [cel("EVIDENTIRAO", r.prijavitelj), "",
         cel("ROK RJEŠAVANJA", rok), "",
         cel("KATEGORIJA / STATUS", f"{kat}  |  {r.status_display}" if kat else r.status_display), ""],
    ]
    id_spans = [
        ("SPAN",(0,0),(1,0)),("SPAN",(2,0),(3,0)),("SPAN",(4,0),(5,0)),
        ("SPAN",(0,1),(1,1)),("SPAN",(2,1),(3,1)),("SPAN",(4,1),(5,1)),
    ]
    if r.naziv_proizvoda or r.broj_radnog_naloga:
        n = len(id_rows)
        id_rows.append([cel("NAZIV I OZNAKA PROIZVODA", r.naziv_proizvoda),"","","",
                         cel("BROJ RADNOG NALOGA", r.broj_radnog_naloga),""])
        id_spans += [("SPAN",(0,n),(3,n)),("SPAN",(4,n),(5,n))]
    if r.stroj or r.osoblje:
        n = len(id_rows)
        id_rows.append([cel("STROJ", r.stroj),"",cel("OSOBLJE", r.osoblje),"","",""])
        id_spans += [("SPAN",(0,n),(1,n)),("SPAN",(2,n),(5,n))]
    if r.kupac_dobavljac or r.referentni_broj:
        n = len(id_rows)
        id_rows.append([cel("KUPAC / DOBAVLJAČ", r.kupac_dobavljac),"","","",
                         cel("REFERENTNI BROJ", r.referentni_broj),""])
        id_spans += [("SPAN",(0,n),(3,n)),("SPAN",(4,n),(5,n))]
    if r.datum_zatvaranja:
        n = len(id_rows)
        id_rows.append([cel("DATUM ZATVARANJA", r.datum_zatvaranja.strftime("%d.%m.%Y %H:%M")),"","","",cel("",""),""])
        id_spans += [("SPAN",(0,n),(3,n)),("SPAN",(4,n),(5,n))]

    id_t = Table(id_rows, colWidths=[col6]*6)
    id_t.setStyle(TableStyle(CP + id_spans))
    story.append(id_t)
    story.append(Spacer(1, 0.3*cm))

    # Opis
    story.append(sek("OPIS NESUKLADNOSTI"))
    opis_t = Table([[Paragraph(_x(r.opis).replace("\n","<br/>"),
                               ParagraphStyle("OP", fontName=_FONT, fontSize=9,
                                              textColor=_TAMNO, leading=14))]],
                   colWidths=[CW])
    opis_t.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
        ("BACKGROUND",(0,0),(-1,-1),_SIVA_BG),
        ("TOPPADDING",(0,0),(-1,-1),9),("BOTTOMPADDING",(0,0),(-1,-1),9),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
    ]))
    story.append(opis_t)

    for label, val in [("KOREKCIJA (NEPOSREDNA RADNJA)", r.korekcija), ("NAPOMENA", r.napomena)]:
        if val:
            t = Table([[cel(label, val)]], colWidths=[CW])
            t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
                                    ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),7),
                                    ("LEFTPADDING",(0,0),(-1,-1),9),("RIGHTPADDING",(0,0),(-1,-1),9)]))
            story.append(t)

    if r.vezana_nesukladnost:
        t = Table([[cel("ISTA NESUKLADNOST EVIDENTIRANA POD BR.", r.vezana_nesukladnost, bold=True)]], colWidths=[CW])
        t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
                                ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),7),
                                ("LEFTPADDING",(0,0),(-1,-1),9)]))
        story.append(t)

    story.append(Spacer(1, 0.3*cm))

    # Analiza uzroka
    if r.analiza_uzroka or r.uzrok_kategorija:
        story.append(sek("ANALIZA UZROKA — METODA 5 ZAŠTO"))
        sAL = ParagraphStyle("AL", fontName=_FONTB, fontSize=7.5, textColor=_SIVA_TL, leading=11)
        sAV = ParagraphStyle("AV", fontName=_FONT, fontSize=9, textColor=_TAMNO, leading=13)
        arows = []
        if r.uzrok_kategorija:
            arows.append([Paragraph("KATEGORIJA", sAL),
                           Paragraph(_x(r.uzrok_kategorija),
                                     ParagraphStyle("KU", fontName=_FONTB, fontSize=9, textColor=_TAMNO, leading=13))])
        if r.analiza_uzroka:
            for i, ln in enumerate([l for l in r.analiza_uzroka.strip().split("\n") if l.strip()]):
                arows.append([Paragraph(f"ZAŠTO {i+1}" if i < 5 else "", sAL),
                               Paragraph(_x(ln.strip()), sAV)])
        if arows:
            a_t = Table(arows, colWidths=[2.6*cm, CW-2.6*cm])
            a_t.setStyle(TableStyle([
                ("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
                ("INNERGRID",(0,0),(-1,-1),0.3,_SIVA_LN),
                ("ROWBACKGROUNDS",(0,0),(-1,-1),[_BIJELA,_SIVA_BG]),
                ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),6),
                ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
                ("VALIGN",(0,0),(-1,-1),"TOP"),
            ]))
            story.append(a_t)
        story.append(Spacer(1, 0.3*cm))

    # CAPA
    if r.capa:
        story.append(sek("KOREKTIVNE I PREVENTIVNE MJERE (CAPA)"))
        sTH = ParagraphStyle("TH", fontName=_FONTB, fontSize=7.5, textColor=_BIJELA, leading=10)
        sTD = ParagraphStyle("TD", fontName=_FONT, fontSize=8.5, textColor=_TAMNO, leading=12)
        c1,c3,c4,c5,c6 = 2.2*cm, 3.0*cm, 2.0*cm, 2.0*cm, 2.8*cm
        c2 = CW - c1 - c3 - c4 - c5 - c6
        capa_data = [[Paragraph("VRSTA",sTH), Paragraph("OPIS MJERE",sTH),
                      Paragraph("ODG. OSOBA",sTH), Paragraph("ROK",sTH),
                      Paragraph("STATUS",sTH), Paragraph("PROVJERA",sTH)]]
        for c in r.capa:
            provjera = _x(c.provjerio) if c.provjerio else ""
            if c.provjerio and c.datum_provjere:
                provjera += f"<br/>{c.datum_provjere.strftime('%d.%m.%Y')}"
            capa_data.append([
                Paragraph(_x(c.vrsta_display), sTD),
                Paragraph(_x(c.opis_mjere).replace("\n","<br/>"), sTD),
                Paragraph(_x(c.odgovorna_osoba), sTD),
                Paragraph(c.rok_izvrsenja.strftime("%d.%m.%Y") if c.rok_izvrsenja else "—", sTD),
                Paragraph(_x(c.status_display), sTD),
                Paragraph(provjera, sTD),
            ])
        ts = [
            ("BACKGROUND",(0,0),(-1,0),_TAMNO),
            ("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
            ("INNERGRID",(0,0),(-1,-1),0.3,_SIVA_LN),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),7),("RIGHTPADDING",(0,0),(-1,-1),5),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[_BIJELA,_SIVA_BG]),
        ]
        for i, c in enumerate(r.capa, 1):
            if c.status == "IZVRSENA":
                ts.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F0FDF4")))
            elif c.je_prekoracen:
                ts.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#FFF1F2")))
        capa_t = Table(capa_data, colWidths=[c1,c2,c3,c4,c5,c6], repeatRows=1)
        capa_t.setStyle(TableStyle(ts))
        story.append(capa_t)
        story.append(Spacer(1, 0.3*cm))

    # Promjene sustava
    if r.promjene_sustava:
        story.append(sek("PROMJENE U SUSTAVU UPRAVLJANJA"))
        ps_val = "Da" if r.promjene_sustava == "DA" else "Ne"
        ps_extra = f"  —  Broj promjene (OB-21): {r.broj_promjene}" if r.promjene_sustava == "DA" and r.broj_promjene else ""
        t = Table([[cel("POTREBNE PROMJENE U SUSTAVU?", ps_val + ps_extra, bold=(r.promjene_sustava=="DA"))]], colWidths=[CW])
        t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
                                ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),8),
                                ("LEFTPADDING",(0,0),(-1,-1),9)]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    # Potpisi
    story.append(Spacer(1, 0.4*cm))
    story.append(sek("POTPISI"))
    spl = ParagraphStyle("PL2", fontName=_FONTB, fontSize=7.5, textColor=_SIVA_TL, leading=11)
    spv = ParagraphStyle("PV2", fontName=_FONT, fontSize=9, textColor=_TAMNO, leading=13)
    spp = ParagraphStyle("PP2", fontName=_FONT, fontSize=8.5, textColor=colors.HexColor("#CBD5E1"), leading=22)
    half = CW / 2
    pot_t = Table([
        [Paragraph("PRIJAVITELJ", spl), Paragraph("VODITELJ KVALITETE", spl)],
        [Paragraph(_x(r.prijavitelj), spv), Paragraph("", spv)],
        [Paragraph("Potpis: _______________________________", spp),
         Paragraph("Potpis: _______________________________", spp)],
        [Paragraph(f"Datum: {r.datum_prijave.strftime('%d.%m.%Y')}", spl),
         Paragraph("Datum: __________________________", spl)],
    ], colWidths=[half, half])
    pot_t.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,_SIVA_LN),
        ("INNERGRID",(0,0),(-1,-1),0.3,_SIVA_LN),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),8),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    story.append(pot_t)

    doc.build(story, onFirstPage=zaglavlje_podnozje, onLaterPages=zaglavlje_podnozje)
    buffer.seek(0)
    return buffer


def generiraj_excel(reklamacije: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reklamacije"

    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(fill_type="solid", fgColor="1A2942")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    stupci = [
        ("Broj predmeta",14),("Naslov",40),("Vrsta",22),("Prioritet",12),
        ("Status",16),("Prijavitelj",20),("Kupac/Dobavljač",25),
        ("Ref. broj",16),("Datum prijave",16),("Rok rješavanja",16),
        ("Datum zatvaranja",18),("Kategorija uzroka",22),
        ("Analiza uzroka",45),("Broj CAPA",10),
    ]
    for col,(naziv,sirina) in enumerate(stupci,1):
        cell = ws.cell(row=1, column=col, value=naziv)
        cell.font = hf; cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = sirina
    ws.row_dimensions[1].height = 30

    STATUS_BOJE = {
        "NOVO":"CCE5FF","U_OBRADI":"FFF3CD","CEKA":"D1ECF1",
        "RIJESENO":"D4EDDA","ZATVORENO":"E2E3E5",
    }
    for row_idx, r in enumerate(reklamacije, 2):
        vrijednosti = [
            r.broj_predmeta, r.naslov, r.vrsta_display, r.prioritet_display,
            r.status_display, r.prijavitelj, r.kupac_dobavljac or "",
            r.referentni_broj or "",
            r.datum_prijave.strftime("%d.%m.%Y %H:%M") if r.datum_prijave else "",
            r.rok_rjesavanja.strftime("%d.%m.%Y") if r.rok_rjesavanja else "",
            r.datum_zatvaranja.strftime("%d.%m.%Y %H:%M") if r.datum_zatvaranja else "",
            r.uzrok_kategorija or "", r.analiza_uzroka or "", r.broj_capa,
        ]
        boja = STATUS_BOJE.get(r.status, "FFFFFF")
        fill = PatternFill(fill_type="solid", fgColor=boja)
        for col, val in enumerate(vrijednosti, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(size=9); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [2,13]))
            if col == 5: cell.fill = fill
        ws.row_dimensions[row_idx].height = 15

    ws2 = wb.create_sheet("CAPA mjere")
    capa_zaglavlja = ["Reklamacija","Vrsta","Opis mjere","Odgovorna osoba","Rok","Status","Datum izvršenja","Rezultat"]
    for col,naziv in enumerate(capa_zaglavlja,1):
        cell = ws2.cell(row=1,column=col,value=naziv)
        cell.font = hf; cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal="center",vertical="center")
    sirine2 = [14,16,40,20,14,14,16,35]
    for i,s in enumerate(sirine2,1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = s

    row2 = 2
    for r in reklamacije:
        for c in r.capa:
            for col,val in enumerate([
                r.broj_predmeta, c.vrsta_display, c.opis_mjere, c.odgovorna_osoba,
                c.rok_izvrsenja.strftime("%d.%m.%Y") if c.rok_izvrsenja else "",
                c.status_display,
                c.datum_izvrsenja.strftime("%d.%m.%Y") if c.datum_izvrsenja else "",
                c.rezultat or "",
            ],1):
                cell = ws2.cell(row=row2,column=col,value=val)
                cell.font = Font(size=9); cell.border = border
                cell.alignment = Alignment(vertical="top",wrap_text=(col in [3,8]))
            row2 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
